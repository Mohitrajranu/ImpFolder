import json
import pandas as pd
from openpyxl import load_workbook
import xlwings as xl

json_file = input("Kindly give complete path of json file(with extension): ")
excel_file = input("Kindly give complete path of excel file(with extension): ")

jdata = pd.read_json(json_file)
df = pd.DataFrame(jdata)

df = df.drop(['Rule_Engine', 'user_name' , 'project_name'], axis=1)
indexNamesArr = df.index.values

for i in range(0 , len(indexNamesArr)):
    indexNamesArr[i] = int(indexNamesArr[i].split("_")[-1])
    
df1=df.sort_index()

df2 = df1.T


wb = load_workbook(filename = excel_file)
ws = wb.active

c=1
for i in range(0 ,len(indexNamesArr)):
    ws.cell(row=3 , column=c , value = str(df2[i][0]))
    c=c+1

wb.save(excel_file)
wb.close()

print("Task Successfully Completed ")


import pyexcel as pe
reader = pe.Reader(excel_file)
HLStart = reader[14 ,3]
HLEnd = reader[14 , 4]

print("HLStart : " ,HLStart)
print("HLEnd   : " ,HLEnd)